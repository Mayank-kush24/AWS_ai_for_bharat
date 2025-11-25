"""
Import utilities for AWS AI for Bharat Tracking System
Handles XLSX parsing, validation, and data mapping
"""
import re
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any
import openpyxl
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


# ============================================
# Constants
# ============================================

WORKSHOP_NAMES = {
    1: "Workshop 1",
    2: "Workshop 2",
    3: "Workshop 3",
    4: "Workshop 4",
    5: "Workshop 5",
    6: "Workshop 6"
}

# Expected headers for Form sheets
FORM_HEADERS = [
    "Team Name",
    "Leader Name",
    "Leader Email",
    "Leader Phone",
    "Team Size",
    "Problem Statements",
    "Name",
    "Email",
    "Book your slot for workshop 1 - Building a Simple Content Summarizer with Amazon Bedrock",
    "Created At",
    "Created By Name",
    "Created By Email",
    "Updated At",
    "Updated By Name",
    "Updated By Email"
]

# Expected headers for Project Submission sheets
PROJECT_HEADERS = [
    "Team Name",
    "Leader Name",
    "Leader Email",
    "Leader Phone",
    "Team Size",
    "Problem Statements",
    "Hands-on Lab Completion Proof",
    "AWS Builder Center Blog Link",
    "Created At",
    "Created By Name",
    "Created By Email",
    "Updated At",
    "Updated By Name",
    "Updated By Email"
]

# Sheet sequence: 1=Project, 2=Form, 3=Form, 4=Project, 5=Form, 6=Project, etc.
SHEET_SEQUENCE = [
    ("project", 1),  # Sheet 1
    ("form", 1),     # Sheet 2
    ("form", 2),     # Sheet 3
    ("project", 2),  # Sheet 4
    ("form", 3),     # Sheet 5
    ("project", 3),  # Sheet 6
    ("form", 4),     # Sheet 7
    ("project", 4),  # Sheet 8
    ("form", 5),     # Sheet 9
    ("project", 5),  # Sheet 10
    ("form", 6),     # Sheet 11
    ("project", 6),  # Sheet 12
]


# ============================================
# Validation Functions
# ============================================

def validate_email(email: str) -> bool:
    """Validate email format"""
    if not email or not isinstance(email, str):
        return False
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email.strip()))


def validate_phone(phone: str) -> Optional[str]:
    """Validate and normalize phone number"""
    if not phone:
        return None
    # Remove all non-digit characters
    digits = re.sub(r'\D', '', str(phone))
    if len(digits) >= 10:
        return digits
    return None


def parse_datetime(date_str: Any) -> Optional[datetime]:
    """Parse datetime from various formats"""
    if not date_str:
        return None
    
    # Handle datetime objects directly
    if isinstance(date_str, datetime):
        return date_str
    
    # Handle date objects
    from datetime import date
    if isinstance(date_str, date):
        return datetime.combine(date_str, datetime.min.time())
    
    if isinstance(date_str, str):
        date_str = date_str.strip()
        if not date_str:
            return None
        
        # Handle ISO 8601 format (e.g., "2025-09-12T11:52:26.894Z" or "2025-09-12T11:52:26Z")
        # This format is common in modern APIs and databases
        # IMPORTANT: We parse UTC times (ending with Z) as naive datetime without timezone conversion
        # The database will store the exact date/time values as provided
        if 'T' in date_str:
            try:
                # Remove 'Z' (UTC indicator) - we treat it as naive datetime, no conversion
                clean_str = date_str.replace('Z', '').replace('z', '')
                
                # Remove timezone offset if present (e.g., "+05:30" or "-05:00")
                # We don't apply timezone conversion, just remove the offset indicator
                tz_pattern = r'[+-]\d{2}:?\d{2}$'
                clean_str = re.sub(tz_pattern, '', clean_str)
                
                # Handle milliseconds/microseconds - preserve them correctly
                microseconds = 0
                if '.' in clean_str:
                    # Split by 'T' to separate date and time
                    if 'T' in clean_str:
                        date_part, time_part = clean_str.split('T', 1)
                        if '.' in time_part:
                            # Extract fractional seconds (can be milliseconds or microseconds)
                            time_base, fractional = time_part.split('.', 1)
                            # Remove any non-digit characters (like Z that might remain)
                            fractional = re.sub(r'\D', '', fractional)
                            # Convert to microseconds (max 6 digits)
                            # If 3 digits (milliseconds), pad to 6 (microseconds)
                            # If 6 digits (microseconds), use as-is
                            # If more than 6, truncate to 6
                            if len(fractional) > 6:
                                fractional = fractional[:6]
                            elif len(fractional) < 6:
                                # Pad with zeros to make it microseconds
                                fractional = fractional.ljust(6, '0')
                            microseconds = int(fractional)
                            time_part = time_base
                        clean_str = f"{date_part}T{time_part}"
                
                # Parse as YYYY-MM-DDTHH:MM:SS (naive datetime, no timezone)
                dt = datetime.strptime(clean_str, '%Y-%m-%dT%H:%M:%S')
                
                # Add microseconds if present
                if microseconds > 0:
                    dt = dt.replace(microsecond=microseconds)
                
                # Return naive datetime - database will store exactly as: YYYY-MM-DD HH:MM:SS.microseconds
                return dt
            except Exception as e:
                # If ISO parsing fails, continue to try other formats
                # Log the error for debugging (can be removed in production)
                logger.debug(f"ISO datetime parsing failed for '{date_str}': {e}")
                pass
        
        # Handle time slot format: "25 Nov, 4:00 - 7:00 PM" or "25 Nov, 4:00 PM - 7:00 PM"
        # Also handle formats without AM/PM: "25 Nov, 4:00 - 7:00"
        # Extract date and start time
        if ',' in date_str and (' - ' in date_str or '-' in date_str):
            try:
                # Split by comma to get date and time parts
                parts = date_str.split(',', 1)
                if len(parts) == 2:
                    date_part = parts[0].strip()  # "25 Nov"
                    time_part = parts[1].strip()  # "4:00 - 7:00 PM"
                    
                    # Extract both start and end time (full range)
                    if ' - ' in time_part:
                        time_parts = time_part.split(' - ')
                        start_time_str = time_parts[0].strip()  # "4:00" or "4:00 PM"
                        end_time_str = time_parts[1].strip() if len(time_parts) > 1 else None  # "7:00 PM"
                    elif '-' in time_part and not time_part.startswith('-'):
                        # Handle single dash without spaces
                        time_parts = time_part.split('-')
                        start_time_str = time_parts[0].strip()
                        end_time_str = time_parts[1].strip() if len(time_parts) > 1 else None
                    else:
                        start_time_str = time_part.strip()
                        end_time_str = None
                    
                    # Check if PM/AM is in the end time but not in start time
                    # Example: "4:00 - 7:00 PM" means both are PM
                    has_pm = 'PM' in time_part.upper()
                    has_am = 'AM' in time_part.upper()
                    
                    # If PM/AM is only in end time, apply it to start time too
                    if not ('PM' in start_time_str.upper() or 'AM' in start_time_str.upper()):
                        if has_pm:
                            start_time_str += ' PM'
                        elif has_am:
                            start_time_str += ' AM'
                    
                    # Parse the date part (day + month abbreviation)
                    # Try to parse "25 Nov" format
                    try:
                        # Get current year or use a default
                        current_year = datetime.now().year
                        # Try parsing with current year
                        date_obj = datetime.strptime(f"{date_part} {current_year}", "%d %b %Y")
                    except:
                        # Try with full month name
                        try:
                            date_obj = datetime.strptime(f"{date_part} {current_year}", "%d %B %Y")
                        except:
                            # If that fails, try without year (will use current year)
                            date_obj = datetime.strptime(date_part, "%d %b")
                            date_obj = date_obj.replace(year=current_year)
                    
                    # Parse the start time part
                    # Handle formats like "4:00 PM", "4:00", "16:00"
                    time_formats = [
                        '%I:%M %p',  # "4:00 PM"
                        '%I:%M%p',   # "4:00PM"
                        '%H:%M',     # "16:00" or "4:00"
                    ]
                    
                    time_obj = None
                    for fmt in time_formats:
                        try:
                            time_obj = datetime.strptime(start_time_str, fmt).time()
                            break
                        except:
                            continue
                    
                    # If time parsing failed, try to extract hour and minute manually
                    if time_obj is None:
                        # Extract numbers from time string
                        time_match = re.search(r'(\d{1,2}):(\d{2})', start_time_str)
                        if time_match:
                            hour = int(time_match.group(1))
                            minute = int(time_match.group(2))
                            # Check if PM - look in original time_part for PM/AM
                            if has_pm and hour < 12:
                                hour += 12
                            elif has_am and hour == 12:
                                hour = 0
                            time_obj = datetime.strptime(f"{hour:02d}:{minute:02d}", "%H:%M").time()
                    
                    if time_obj:
                        # For now, return the start time as datetime
                        # We'll store the full range as a string in the database
                        # But for datetime field, we use start time
                        return datetime.combine(date_obj.date(), time_obj)
            except Exception as e:
                # If parsing fails, continue to try other formats
                pass
        
        # Try common formats
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%d',
            '%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y',
            '%m/%d/%Y %H:%M:%S',
            '%m/%d/%Y',
            '%d %b %Y %H:%M:%S',  # "25 Nov 2024 14:00:00"
            '%d %b %Y %H:%M',      # "25 Nov 2024 14:00"
            '%d %B %Y %H:%M:%S',   # "25 November 2024 14:00:00"
            '%d %B %Y %H:%M',      # "25 November 2024 14:00"
        ]
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except:
                continue
        
        # Try parsing Excel date serial numbers
        try:
            if isinstance(date_str, (int, float)):
                # Excel serial date (days since 1900-01-01)
                from datetime import timedelta
                excel_epoch = datetime(1900, 1, 1)
                return excel_epoch + timedelta(days=int(date_str) - 2)  # Excel epoch is off by 2 days
        except:
            pass
    
    return None


def coerce_boolean(value: Any) -> bool:
    """Coerce value to boolean"""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ('true', '1', 'yes', 'y', 'on')
    if isinstance(value, (int, float)):
        return bool(value)
    return False


def normalize_string(value: Any) -> Optional[str]:
    """Normalize string values"""
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned else None
    return str(value).strip() if str(value).strip() else None


# ============================================
# XLSX Parsing Functions
# ============================================

def read_xlsx_file(file_path: str) -> openpyxl.Workbook:
    """Read XLSX file and return workbook"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        return workbook
    except Exception as e:
        raise ValueError(f"Failed to read XLSX file: {str(e)}")


def get_sheet_headers(sheet, max_row: int = 1) -> List[str]:
    """Extract headers from first row of sheet"""
    headers = []
    for cell in sheet[1]:
        value = cell.value
        headers.append(str(value).strip() if value else "")
    return headers


def validate_sheet_headers(headers: List[str], expected_type: str) -> Tuple[bool, List[str]]:
    """Validate sheet headers against expected format"""
    errors = []
    
    if expected_type == "form":
        expected = FORM_HEADERS
    elif expected_type == "project":
        expected = PROJECT_HEADERS
    else:
        return False, ["Unknown sheet type"]
    
    # Check if we have enough columns
    if len(headers) < len(expected):
        errors.append(f"Expected at least {len(expected)} columns, got {len(headers)}")
    
    # Check for critical headers (flexible matching)
    critical_headers = {
        "form": ["Email", "Name", "Team Name"],
        "project": ["Email", "Name", "Team Name", "AWS Builder Center Blog Link"]
    }
    
    header_lower = [h.lower() for h in headers]
    for critical in critical_headers.get(expected_type, []):
        if not any(critical.lower() in h.lower() for h in headers):
            errors.append(f"Missing critical header: {critical}")
    
    return len(errors) == 0, errors


def find_column_index(headers: List[str], search_terms: List[str]) -> Optional[int]:
    """Find column index by searching for terms in headers"""
    for idx, header in enumerate(headers):
        header_lower = header.lower()
        for term in search_terms:
            if term.lower() in header_lower:
                return idx
    return None


def parse_form_sheet(sheet, workshop_num: int) -> Tuple[List[Dict], List[str]]:
    """Parse a form sheet and return list of records and errors"""
    records = []
    errors = []
    
    # Get headers
    headers = get_sheet_headers(sheet)
    is_valid, validation_errors = validate_sheet_headers(headers, "form")
    if not is_valid:
        errors.extend(validation_errors)
        return records, errors
    
    # Find column indices
    name_idx = find_column_index(headers, ["Name"])
    email_idx = find_column_index(headers, ["Email"])
    team_name_idx = find_column_index(headers, ["Team Name", "TeamName"])
    time_slot_idx = find_column_index(headers, ["Book your slot", "time slot", "Time Slot"])
    created_at_idx = find_column_index(headers, ["Created At", "CreatedAt"])
    updated_at_idx = find_column_index(headers, ["Updated At", "UpdatedAt"])
    
    if name_idx is None or email_idx is None:
        errors.append("Missing required columns: Name or Email")
        return records, errors
    
    workshop_name = WORKSHOP_NAMES.get(workshop_num, f"Workshop {workshop_num}")
    
    # Process rows (skip header row)
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Skip empty rows
        if not any(cell.value for cell in row):
            continue
        
        try:
            name = normalize_string(row[name_idx].value) if name_idx is not None else None
            email = normalize_string(row[email_idx].value) if email_idx is not None else None
            
            if not name or not email:
                errors.append(f"Row {row_num}: Missing name or email")
                continue
            
            # Validate email
            if not validate_email(email):
                errors.append(f"Row {row_num}: Invalid email format: {email}")
                continue
            
            # Extract time slot
            time_slot = None
            if time_slot_idx is not None and row[time_slot_idx].value:
                time_slot = parse_datetime(row[time_slot_idx].value)
            
            # Extract timestamps
            created_at = None
            if created_at_idx is not None and row[created_at_idx].value:
                created_at = parse_datetime(row[created_at_idx].value)
            
            updated_at = None
            if updated_at_idx is not None and row[updated_at_idx].value:
                updated_at = parse_datetime(row[updated_at_idx].value)
            
            # Extract team name
            team_name = None
            if team_name_idx is not None and row[team_name_idx].value:
                team_name = normalize_string(row[team_name_idx].value)
            
            record = {
                "workshop_name": workshop_name,
                "email": email.lower().strip(),
                "name": name,
                "time_slot": time_slot,
                "created_at": created_at or datetime.now(),
                "updated_at": updated_at or datetime.now(),
                "team_name": team_name,
                "row_number": row_num
            }
            
            records.append(record)
            
        except Exception as e:
            errors.append(f"Row {row_num}: Error parsing row - {str(e)}")
            continue
    
    return records, errors


def parse_project_sheet(sheet, workshop_num: int) -> Tuple[List[Dict], List[str]]:
    """Parse a project submission sheet and return list of records and errors"""
    records = []
    errors = []
    
    # Get headers
    headers = get_sheet_headers(sheet)
    is_valid, validation_errors = validate_sheet_headers(headers, "project")
    if not is_valid:
        errors.extend(validation_errors)
        return records, errors
    
    # Find column indices
    name_idx = find_column_index(headers, ["Name"])
    email_idx = find_column_index(headers, ["Email"])
    team_name_idx = find_column_index(headers, ["Team Name", "TeamName"])
    project_link_idx = find_column_index(headers, ["AWS Builder Center Blog Link", "Blog Link", "Project Link"])
    created_at_idx = find_column_index(headers, ["Created At", "CreatedAt"])
    updated_at_idx = find_column_index(headers, ["Updated At", "UpdatedAt"])
    
    if name_idx is None or email_idx is None:
        errors.append("Missing required columns: Name or Email")
        return records, errors
    
    workshop_name = WORKSHOP_NAMES.get(workshop_num, f"Workshop {workshop_num}")
    
    # Process rows (skip header row)
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Skip empty rows
        if not any(cell.value for cell in row):
            continue
        
        try:
            name = normalize_string(row[name_idx].value) if name_idx is not None else None
            email = normalize_string(row[email_idx].value) if email_idx is not None else None
            
            if not name or not email:
                errors.append(f"Row {row_num}: Missing name or email")
                continue
            
            # Validate email
            if not validate_email(email):
                errors.append(f"Row {row_num}: Invalid email format: {email}")
                continue
            
            # Extract project link
            project_link = None
            if project_link_idx is not None and row[project_link_idx].value:
                project_link = normalize_string(row[project_link_idx].value)
            
            # Extract timestamps
            created_at = None
            if created_at_idx is not None and row[created_at_idx].value:
                created_at = parse_datetime(row[created_at_idx].value)
            
            updated_at = None
            if updated_at_idx is not None and row[updated_at_idx].value:
                updated_at = parse_datetime(row[updated_at_idx].value)
            
            # Extract team name
            team_name = None
            if team_name_idx is not None and row[team_name_idx].value:
                team_name = normalize_string(row[team_name_idx].value)
            
            record = {
                "workshop_name": workshop_name,
                "email": email.lower().strip(),
                "name": name,
                "project_link": project_link,
                "valid": False,  # Default to False, can be updated later
                "team_id": team_name,  # Using team_name as team_id
                "created_at": created_at or datetime.now(),
                "updated_at": updated_at or datetime.now(),
                "row_number": row_num
            }
            
            records.append(record)
            
        except Exception as e:
            errors.append(f"Row {row_num}: Error parsing row - {str(e)}")
            continue
    
    return records, errors


def parse_user_pii_sheet(sheet) -> Tuple[List[Dict], List[str]]:
    """Parse User PII sheet and return list of records and errors"""
    records = []
    errors = []
    
    # Get headers
    headers = get_sheet_headers(sheet)
    
    # Find column indices (flexible matching)
    email_idx = find_column_index(headers, ["Email", "email"])
    name_idx = find_column_index(headers, ["Name", "name"])
    phone_idx = find_column_index(headers, ["Phone", "phone", "Phone Number"])
    gender_idx = find_column_index(headers, ["Gender", "gender"])
    country_idx = find_column_index(headers, ["Country", "country"])
    state_idx = find_column_index(headers, ["State", "state"])
    city_idx = find_column_index(headers, ["City", "city"])
    dob_idx = find_column_index(headers, ["Date of Birth", "DOB", "dob", "Birth Date"])
    designation_idx = find_column_index(headers, ["Designation", "designation"])
    class_stream_idx = find_column_index(headers, ["Class", "Stream", "Class/Stream"])
    degree_idx = find_column_index(headers, ["Degree", "Passout Year", "Passout"])
    occupation_idx = find_column_index(headers, ["Occupation", "occupation"])
    linkedin_idx = find_column_index(headers, ["LinkedIn", "linkedin", "LinkedIn URL"])
    academy_idx = find_column_index(headers, ["Academy", "Participated"])
    registration_idx = find_column_index(headers, ["Registration", "Registration Date"])
    
    if email_idx is None or name_idx is None:
        errors.append("Missing required columns: Email or Name")
        return records, errors
    
    # Process rows (skip header row)
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Skip empty rows
        if not any(cell.value for cell in row):
            continue
        
        try:
            email = normalize_string(row[email_idx].value) if email_idx is not None else None
            name = normalize_string(row[name_idx].value) if name_idx is not None else None
            
            if not email or not name:
                errors.append(f"Row {row_num}: Missing email or name")
                continue
            
            # Validate email
            if not validate_email(email):
                errors.append(f"Row {row_num}: Invalid email format: {email}")
                continue
            
            # Extract phone
            phone = None
            if phone_idx is not None and row[phone_idx].value:
                phone = validate_phone(str(row[phone_idx].value))
            
            # Extract date of birth
            date_of_birth = None
            if dob_idx is not None and row[dob_idx].value:
                dob = parse_datetime(row[dob_idx].value)
                if dob:
                    date_of_birth = dob.date() if hasattr(dob, 'date') else dob
            
            # Extract registration datetime
            registration_date_time = datetime.now()
            if registration_idx is not None and row[registration_idx].value:
                reg_date = parse_datetime(row[registration_idx].value)
                if reg_date:
                    registration_date_time = reg_date
            
            # Extract degree passout year
            degree_passout_year = None
            if degree_idx is not None and row[degree_idx].value:
                try:
                    degree_value = str(row[degree_idx].value)
                    # Try to extract year from string
                    year_match = re.search(r'\d{4}', degree_value)
                    if year_match:
                        degree_passout_year = int(year_match.group())
                except:
                    pass
            
            record = {
                "email": email.lower().strip(),
                "name": name,
                "phone_number": phone,
                "gender": normalize_string(row[gender_idx].value) if gender_idx is not None else None,
                "country": normalize_string(row[country_idx].value) if country_idx is not None else None,
                "state": normalize_string(row[state_idx].value) if state_idx is not None else None,
                "city": normalize_string(row[city_idx].value) if city_idx is not None else None,
                "date_of_birth": date_of_birth,
                "designation": normalize_string(row[designation_idx].value) if designation_idx is not None else None,
                "class_stream": normalize_string(row[class_stream_idx].value) if class_stream_idx is not None else None,
                "degree_passout_year": degree_passout_year,
                "occupation": normalize_string(row[occupation_idx].value) if occupation_idx is not None else None,
                "linkedin": normalize_string(row[linkedin_idx].value) if linkedin_idx is not None else None,
                "participated_in_academy_1_0": coerce_boolean(row[academy_idx].value) if academy_idx is not None else False,
                "registration_date_time": registration_date_time,
                "row_number": row_num
            }
            
            records.append(record)
            
        except Exception as e:
            errors.append(f"Row {row_num}: Error parsing row - {str(e)}")
            continue
    
    return records, errors


def parse_master_workbook(file_path: str) -> Dict[str, Any]:
    """Parse the master workbook with 12 sheets"""
    workbook = read_xlsx_file(file_path)
    
    result = {
        "sheets": [],
        "total_records": 0,
        "total_errors": [],
        "workshops_processed": []
    }
    
    # Validate sheet count
    if len(workbook.sheetnames) < 12:
        result["total_errors"].append(f"Expected 12 sheets, found {len(workbook.sheetnames)}")
        return result
    
    # Process each sheet according to sequence
    for sheet_idx, (sheet_type, workshop_num) in enumerate(SHEET_SEQUENCE, start=1):
        if sheet_idx > len(workbook.sheetnames):
            result["total_errors"].append(f"Sheet {sheet_idx} missing")
            continue
        
        sheet = workbook[workbook.sheetnames[sheet_idx - 1]]
        sheet_name = workbook.sheetnames[sheet_idx - 1]
        
        sheet_result = {
            "sheet_index": sheet_idx,
            "sheet_name": sheet_name,
            "sheet_type": sheet_type,
            "workshop_num": workshop_num,
            "records": [],
            "errors": []
        }
        
        try:
            if sheet_type == "form":
                records, errors = parse_form_sheet(sheet, workshop_num)
            elif sheet_type == "project":
                records, errors = parse_project_sheet(sheet, workshop_num)
            else:
                errors = [f"Unknown sheet type: {sheet_type}"]
                records = []
            
            sheet_result["records"] = records
            sheet_result["errors"] = errors
            sheet_result["rows_read"] = len(records)
            sheet_result["rows_inserted"] = 0  # Will be updated after DB insertion
            sheet_result["rows_updated"] = 0
            
            result["sheets"].append(sheet_result)
            result["total_records"] += len(records)
            result["total_errors"].extend(errors)
            
            if workshop_num not in result["workshops_processed"]:
                result["workshops_processed"].append(workshop_num)
        
        except Exception as e:
            error_msg = f"Error processing sheet {sheet_idx} ({sheet_name}): {str(e)}"
            sheet_result["errors"].append(error_msg)
            result["total_errors"].append(error_msg)
            result["sheets"].append(sheet_result)
    
    workbook.close()
    return result


def parse_user_pii_workbook(file_path: str) -> Dict[str, Any]:
    """Parse User PII workbook"""
    workbook = read_xlsx_file(file_path)
    
    result = {
        "records": [],
        "errors": [],
        "rows_read": 0,
        "rows_inserted": 0,
        "rows_updated": 0
    }
    
    # Process first sheet (or all sheets if multiple)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        records, errors = parse_user_pii_sheet(sheet)
        result["records"].extend(records)
        result["errors"].extend(errors)
    
    result["rows_read"] = len(result["records"])
    workbook.close()
    return result

